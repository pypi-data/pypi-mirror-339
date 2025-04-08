# !/usr/bin/python3
# -*- coding: utf-8 -*-
"""
功能描述
----------------------------------------------------
@Project :   marmot-xinhou-openai-framework  
@File    :   ExcelUtil.py    
@Contact :   sp_hrz@qq.com

@Modify Time      @Author    @Version    @Desciption
------------      -------    --------    -----------
2022/12/1 15:42   shenpeng   1.0         None
"""

import pandas as pd
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill


class ExcelUtil:
    """
    Excel工具类，用于读写Excel文件
    """

    @staticmethod
    def read_excel(file_path, sheet_name=None):
        """
        读取Excel文件

        :param file_path: Excel文件路径
        :param sheet_name: 工作表名称，如果为None则读取第一个工作表
        :return: 返回pandas DataFrame对象
        """
        return pd.read_excel(file_path, sheet_name=sheet_name)

    @staticmethod
    def write_excel(data, file_path, sheet_name='Sheet1', style_dict=None):
        """
        写入Excel文件

        :param data: pandas DataFrame对象
        :param file_path: Excel文件路径
        :param sheet_name: 工作表名称
        :param style_dict: 表格样式字典
        :return: None
        """
        if style_dict is None:
            style_dict = {}

        writer = pd.ExcelWriter(file_path, engine='openpyxl')
        data.to_excel(writer, sheet_name=sheet_name, index=False)

        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # 设置列宽
        for i, col in enumerate(data.columns):
            column_len = max(len(str(col)), *(len(str(row)) for row in data[col]))
            column_len = min(column_len, 100)
            worksheet.column_dimensions[chr(ord('A') + i)].width = column_len

        # 设置表格样式
        font_name = style_dict.get('font_name', 'Arial')
        font_size = style_dict.get('font_size', 12)
        font_bold = style_dict.get('font_bold', False)
        font_italic = style_dict.get('font_italic', False)
        font_color = style_dict.get('font_color', '000000')
        font = Font(name=font_name, size=font_size, bold=font_bold, italic=font_italic, color=font_color)

        fill_pattern = style_dict.get('fill_pattern', 'solid')
        fill_color = style_dict.get('fill_color', 'FFFFFF')
        fill = PatternFill(patternType=fill_pattern, fgColor=fill_color)

        border_left = style_dict.get('border_left', 'none')
        border_right = style_dict.get('border_right', 'none')
        border_top = style_dict.get('border_top', 'none')
        border_bottom = style_dict.get('border_bottom', 'none')
        border_left_color = style_dict.get('border_left_color', '000000')
        border_right_color = style_dict.get('border_right_color', '000000')
        border_top_color = style_dict.get('border_top_color', '000000')
        border_bottom_color = style_dict.get('border_bottom_color', '000000')
        border = Border(left=Side(border_style=border_left, color=border_left_color),
                        right=Side(border_style=border_right, color=border_right_color),
                        top=Side(border_style=border_top, color=border_top_color),
                        bottom=Side(border_style=border_bottom, color=border_bottom_color))

        align_horizontal = style_dict.get('align_horizontal', 'left')
        align_vertical = style_dict.get('align_vertical', 'center')
        align = Alignment(horizontal=align_horizontal, vertical=align_vertical)

        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = font
                cell.fill = fill
                cell.border = border
                cell.alignment = align
        writer.close()

    @staticmethod
    def read_csv(file_path):
        """
        读取CSV文件

        :param file_path: CSV文件路径
        :return: 返回pandas DataFrame对象
        """
        return pd.read_csv(file_path)

    @staticmethod
    def write_csv(data, file_path):
        """
        写入CSV文件

        :param data: pandas DataFrame对象
        :param file_path: CSV文件路径
        :return: None
        """
        data.to_csv(file_path, index=False)


if __name__ == '__main__':
    data = pd.DataFrame({'测试列A': [1, 2, 3], '测试列B': ['a', 'b', 'c']})
    ExcelUtil.write_excel(data, 'test.xlsx', style_dict={'font_name': 'SimSun', 'font_size': 14, 'font_bold': True})

    data = pd.DataFrame({'测试列A': [1, 2, 3], '测试列B': ['a', 'b', 'c']})
    ExcelUtil.write_csv(data, 'test.csv')
