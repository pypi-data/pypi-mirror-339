import io
from collections.abc import AsyncIterable, Iterable, Mapping
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta, timezone
from typing import (
    Any,
    Callable,
    Optional,
    Union,
)
from xml.etree import ElementTree as ETree

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

DocumentsStream = Iterable[Mapping[str, Any]]
DocumentsStreamAsync = AsyncIterable[Mapping[str, Any]]
ModifyWorkbookCallable = Callable[[Workbook], None]
ModifyCellCallable = Callable[[Cell], None]
ModifyWorksheetCallable = Callable[[Worksheet], None]


@dataclass
class ColumnDescription:
    #: Document field with corresponding value.
    field: str
    #: Column name in output file.
    title: Optional[str] = None
    #: Refer to [this](https://learn.microsoft.com/en-us/dotnet/api/documentformat.openxml.spreadsheet.numberingformat?view=openxml-3.0.1)
    #: list for possible options.
    number_format: str = "General"


ColumnsDescription = list[ColumnDescription]


def _convert_moment_to_excel_string(
    value: Union[datetime, date, time, timedelta],
    display_timezone: timezone,
):
    # Default Excel epoch
    epoch = datetime(1899, 12, 31, tzinfo=None)

    if isinstance(value, datetime):
        value = value.astimezone(display_timezone).replace(tzinfo=None)
        delta = value - epoch
    elif isinstance(value, date):
        value = datetime.combine(value, time.min, tzinfo=timezone.utc)
        value = value.astimezone(display_timezone).replace(tzinfo=None)
        delta = value - epoch
    elif isinstance(value, time):
        value = datetime.combine(epoch, value, tzinfo=timezone.utc)
        value = value.astimezone(display_timezone).replace(tzinfo=None)
        delta = value - epoch
    elif isinstance(value, timedelta):
        delta = value

    excel_time = delta.days + (
        float(delta.seconds) + float(delta.microseconds) / 1e6
    ) / (60 * 60 * 24)

    # Adjust for Excel's incorrect leap year handling in 1900
    if excel_time > 59:
        excel_time += 1

    return str(excel_time)


def convert_cell_value_to_excel_string(
    value: Any,
    display_timezone: timezone,
):
    if value is None:
        return ""
    if isinstance(value, (datetime, date, time, timedelta)):
        return _convert_moment_to_excel_string(value, display_timezone)
    return str(value)


def generate_template(
    columns: ColumnsDescription,
    freeze_panes: Optional[str] = None,
    modify_worksheet: Optional[ModifyWorksheetCallable] = None,
    modify_workbook: Optional[ModifyWorkbookCallable] = None,
) -> tuple[bytes, int]:
    wb = Workbook()
    assert wb.active, "Created template workbook has no active sheet"

    for i, column_description in enumerate(columns):
        header_cell = wb.active.cell(1, i + 1)
        assert isinstance(header_cell, Cell), "header cell is not a cell"
        header_cell.value = column_description.title or column_description.field
        header_cell.font = Font(bold=True)

        data_cell = wb.active.cell(2, i + 1)
        assert isinstance(data_cell, Cell), "header cell is not a cell"
        data_cell.value = "" if column_description.number_format == "General" else 0
        data_cell.number_format = column_description.number_format

    if freeze_panes:
        wb.active.freeze_panes = freeze_panes

    if modify_worksheet:
        modify_worksheet(wb.active)

    if modify_workbook:
        modify_workbook(wb)

    with io.BytesIO() as fh:
        wb.save(fh)
        return (fh.getvalue(), wb.index(wb.active))


def fill_data_row(
    row: ETree.Element,
    columns: ColumnsDescription,
    display_timezone: timezone,
    document: Mapping[str, Any],
    row_number: int,
):
    for i, column_description in enumerate(columns):
        value = convert_cell_value_to_excel_string(
            document.get(column_description.field),
            display_timezone=display_timezone,
        )

        if column_description.number_format == "General":
            if row[i].get("t") != "inlineStr":
                raise Exception(
                    "unexpected data type in template for field "
                    f"{column_description.field}: {row[i].get('t')}"
                )
            row[i].append(ETree.Element("is"))
            row[i][0].append(ETree.Element("t"))
            row[i][0][0].text = value
        else:
            if row[i].find("v") is None:
                row[i].append(ETree.Element("v"))
            row[i][0].text = value

        row[i].attrib["r"] = f"{get_column_letter(i + 1)}{row_number}"
