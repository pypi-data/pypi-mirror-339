from dataclasses import dataclass
from typing import List, Dict, Optional, Tuple
import pandas as pd
from openpyxl import load_workbook
import re
from pathlib import Path

@dataclass
class Table:
    table_id: str
    dataframe: pd.DataFrame
    sheet_name: str
    start_cell: str
    pattern_match: bool

@dataclass
class DocumentAnalysis:
    tables: List[Table]
    key_values: Dict[str, str]

class TableExtractor:
    def __init__(self, pattern: str):
        self.pattern = re.compile(pattern)
        
    def match_table(self, table_data: List[List[str]]) -> bool:
        """Check if table matches the given pattern"""
        # Convert table data to a single string for pattern matching
        table_text = ' '.join([' '.join(str(cell) for cell in row) for row in table_data])
        return bool(self.pattern.search(table_text))
    
    def convert_to_dataframe(self, table_data: List[List[str]]) -> pd.DataFrame:
        """Convert table data to pandas DataFrame"""
        return pd.DataFrame(table_data[1:], columns=table_data[0])

class KeyValueExtractor:
    def __init__(self):
        # Common patterns for key-value pairs
        self.patterns = [
            r'([^:]+):\s*(.+)',  # Key: Value
            r'([^=]+)=\s*(.+)',  # Key=Value
            r'([^:]+)\s*-\s*(.+)'  # Key - Value
        ]
        self.compiled_patterns = [re.compile(p) for p in self.patterns]
    
    def find_key_value_pairs(self, text: str) -> Dict[str, str]:
        """Extract key-value pairs from text"""
        pairs = {}
        for pattern in self.compiled_patterns:
            matches = pattern.findall(text)
            for key, value in matches:
                pairs[key.strip()] = value.strip()
        return pairs

class ExcelDocumentHandler:
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.workbook = None
        self.table_extractor = None
        self.key_value_extractor = KeyValueExtractor()
        
    def load_document(self) -> None:
        """Load and validate Excel document"""
        if not self.file_path.exists():
            raise FileNotFoundError(f"Document not found: {self.file_path}")
        if self.file_path.suffix.lower() != '.xlsx':
            raise ValueError("Only .xlsx files are supported")
        self.workbook = load_workbook(self.file_path, data_only=True)
        
    def set_table_pattern(self, pattern: str) -> None:
        """Set the pattern for table matching"""
        self.table_extractor = TableExtractor(pattern)
        
    def _find_table_boundaries(self, sheet, start_row: int, start_col: int) -> Tuple[int, int]:
        """Find the boundaries of a table starting at the given cell"""
        # Find the last row with data
        last_row = start_row
        while last_row <= sheet.max_row and any(sheet.cell(row=last_row, column=col).value for col in range(start_col, sheet.max_column + 1)):
            last_row += 1
        last_row -= 1
        
        # Find the last column with data
        last_col = start_col
        while last_col <= sheet.max_column and any(sheet.cell(row=row, column=last_col).value for row in range(start_row, last_row + 1)):
            last_col += 1
        last_col -= 1
        
        return last_row, last_col
        
    def extract_tables(self) -> List[Table]:
        """Extract all tables from document"""
        if not self.workbook:
            raise ValueError("Document not loaded. Call load_document() first.")
        if not self.table_extractor:
            raise ValueError("Table pattern not set. Call set_table_pattern() first.")
            
        tables = []
        table_id = 1
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            
            # Scan the sheet for potential tables
            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    if cell.value:  # Found a non-empty cell
                        # Check if this is the start of a table
                        last_row, last_col = self._find_table_boundaries(sheet, row, col)
                        
                        # Extract table data
                        table_data = []
                        for r in range(row, last_row + 1):
                            row_data = []
                            for c in range(col, last_col + 1):
                                cell_value = sheet.cell(row=r, column=c).value
                                row_data.append(str(cell_value) if cell_value is not None else '')
                            table_data.append(row_data)
                        
                        # Skip if table is empty or too small
                        if len(table_data) < 2 or len(table_data[0]) < 2:
                            continue
                            
                        # Check if table matches pattern
                        matches_pattern = self.table_extractor.match_table(table_data)
                        
                        # Convert to DataFrame
                        try:
                            df = self.table_extractor.convert_to_dataframe(table_data)
                            tables.append(Table(
                                table_id=f"table_{table_id}",
                                dataframe=df,
                                sheet_name=sheet_name,
                                start_cell=f"{chr(64 + col)}{row}",
                                pattern_match=matches_pattern
                            ))
                            table_id += 1
                        except Exception as e:
                            print(f"Error processing table at {sheet_name}!{chr(64 + col)}{row}: {str(e)}")
                            continue
                        
                        # Skip the rest of this table
                        col = last_col
                        break
                
        return tables
    
    def extract_key_values(self) -> Dict[str, str]:
        """Extract key-value pairs from document"""
        if not self.workbook:
            raise ValueError("Document not loaded. Call load_document() first.")
            
        all_pairs = {}
        
        for sheet in self.workbook:
            # Check sheet name for key-value pairs
            pairs = self.key_value_extractor.find_key_value_pairs(sheet.title)
            all_pairs.update(pairs)
            
            # Check cell values for key-value pairs
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        pairs = self.key_value_extractor.find_key_value_pairs(str(cell.value))
                        all_pairs.update(pairs)
            
        return all_pairs
    
    def analyze_document(self) -> DocumentAnalysis:
        """Perform complete document analysis"""
        tables = self.extract_tables()
        key_values = self.extract_key_values()
        return DocumentAnalysis(tables=tables, key_values=key_values) 