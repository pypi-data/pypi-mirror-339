def read_command_line():
    from argparse import ArgumentParser as AP
    from canvas_selector import choose_options
    parser = AP()
    parser.add_argument('-c', '--course', help="Canvas courseID",
                        default=None)
    parser.add_argument('-s', '--semester', help='semester id, e.g. "2241"',
                        default=None)

    args = parser.parse_args()
    args.groups = True
    args = choose_options(args)
    return args 


def load_spreadsheets():
    from importlib.resources import files
    from openpyxl import load_workbook
    template_spreadsheet = [files('PeerMark.data').joinpath(f'team_of_{ii}.xlsx') for ii in [3, 4, 5]]
    return [load_workbook(temp) for temp in template_spreadsheet]


def write_sheet(fname, team, wb):
    # Construct student spreadsheet containing the necessary student names
    from openpyxl import load_workbook
    from openpyxl.styles import Protection
    from canvas_selector import mkdir

    mkdir("tempupload")
    opname = f"tempupload/{fname}_peer_review.xlsx"
    wb.save(filename=opname)

    opwb = load_workbook(opname)
    opws = opwb.active
    opws.title = "Peer Review"
    opws.protection.sheet = True

    for col, name in enumerate(team, 2):
        opws.cell(1, col, name)
        for row in [2, 3, 4, 5, 6, 7, 9, 11]:
            opws.cell(row, col).protection = Protection(locked=False)

    opwb.save(opname)
    return opname


def upload_spreadsheets():
    from tqdm import tqdm
    args = read_command_line()
    wbs = load_spreadsheets()
    groups = [g for g in args.group.get_groups()]
    for group in tqdm(groups, desc="uploading spreadsheets", ascii=True):
        group_members = []
        for user in group.get_users():
            group_members.append(args.course.get_user(user.id).name)

        if group_members:
            num_studs = len(group_members)
            wb = wbs[num_studs - 3]
            tname = group.name.replace(" ", "_")
            fname = write_sheet(tname, group_members, wb)
            group.upload(fname)


def main():
    from canvas_selector import cleanup
    upload_spreadsheets()
    cleanup('tempupload')


if __name__ == '__main__':
    main()