"""
Parser for Microsoft Excel spreadsheets.
"""

import logging
from pathlib import Path

import openpyxl
import pandas as pd

from neuradoc.parsers import BaseParser
from neuradoc.models.element import Element, ElementType

logger = logging.getLogger(__name__)


class Parser(BaseParser):
    """Parser for Microsoft Excel documents."""
    
    def parse(self, file_path):
        """
        Parse an Excel document.
        
        Args:
            file_path (str): Path to the Excel file
            
        Returns:
            dict: Parsed content with metadata
        """
        self.validate_file(file_path)
        
        try:
            # Load the workbook
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # Extract document metadata
            metadata = {
                'title': Path(file_path).stem,
                'sheet_names': workbook.sheetnames,
                'file_path': file_path,
                'file_type': 'xlsx',
                'num_sheets': len(workbook.sheetnames)
            }
            
            # Extract elements (tables, charts, etc.) from each sheet
            elements = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Get the data range
                data_rows = []
                for row in sheet.iter_rows(values_only=True):
                    # Skip completely empty rows
                    if any(cell is not None and str(cell).strip() != '' for cell in row):
                        data_rows.append(row)
                
                if data_rows:
                    # Convert to list of lists, replacing None with empty string
                    table_data = []
                    for row in data_rows:
                        table_data.append(['' if cell is None else str(cell) for cell in row])
                    
                    elements.append(Element(
                        element_type=ElementType.TABLE,
                        content=table_data,
                        metadata={'sheet_name': sheet_name},
                        position={'sheet': sheet_name}
                    ))
                
                # Look for charts and graphs (openpyxl has limited support for this)
                # We can detect the presence of charts but can't extract them easily
                if hasattr(sheet, '_charts') and sheet._charts:
                    for i, chart in enumerate(sheet._charts):
                        elements.append(Element(
                            element_type=ElementType.DIAGRAM,
                            content=f"Chart in sheet '{sheet_name}'",
                            metadata={'chart_type': getattr(chart, 'type', 'unknown')},
                            position={'sheet': sheet_name, 'index': i}
                        ))
                
                # Check for images
                if hasattr(sheet, '_images') and sheet._images:
                    for i, img in enumerate(sheet._images):
                        elements.append(Element(
                            element_type=ElementType.IMAGE,
                            content=img,
                            position={'sheet': sheet_name, 'index': i}
                        ))
            
            return {
                'metadata': metadata,
                'elements': elements
            }
            
        except Exception as e:
            logger.error(f"Error parsing Excel file {file_path}: {e}")
            raise
    
    def _convert_to_dataframe(self, sheet):
        """
        Convert sheet data to pandas DataFrame.
        
        Args:
            sheet: Excel worksheet
            
        Returns:
            DataFrame: Pandas DataFrame containing the sheet data
        """
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        
        if not data:
            return pd.DataFrame()
        
        # Use first row as header if it looks like headers
        if all(isinstance(cell, str) for cell in data[0]):
            return pd.DataFrame(data[1:], columns=data[0])
        else:
            return pd.DataFrame(data)
