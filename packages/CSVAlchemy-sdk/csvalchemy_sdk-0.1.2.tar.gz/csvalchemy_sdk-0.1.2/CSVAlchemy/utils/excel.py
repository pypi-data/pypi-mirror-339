import logging
from typing import Dict, Any, List
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

logger = logging.getLogger(__name__)


def load_workbook_from_file(file_path: str) -> Dict[str, Dict[str, Any]]:
    """
    Load an Excel workbook from a file path into a dictionary format

    Args:
        file_path: Path to the Excel file

    Returns:
        Dictionary representation of the workbook
        {
            'Sheet1': {
                'A1': {'value': 'Header', 'type': 'string', 'format': 'General'},
                'B1': {'value': 10, 'type': 'number', 'format': 'Number'},
                'C1': {'value': '=SUM(A1:B1)', 'type': 'formula', 'format': 'General', 'formula': '=SUM(A1:B1)'}
            },
            'Sheet2': {...}
        }

    Raises:
        ValueError: If the file is not a valid Excel file
    """
    try:
        logger.info(f"Loading workbook from {file_path}")

        # Load workbook
        wb = openpyxl.load_workbook(file_path, data_only=False)

        # Convert to dictionary format
        workbook_dict = {}

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_dict = {}

            # Detect merged cells
            merged_ranges = sheet.merged_cells.ranges
            merged_cells = {}

            for merged_range in merged_ranges:
                # Get the top-left cell of the merged range
                top_left = merged_range.min_row, merged_range.min_col

                # Store all cells in the merged range
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        merged_cells[(row, col)] = top_left

            # Process each cell
            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row, col)

                    # Skip empty cells
                    if (
                        cell.value is None
                        and cell.data_type == "n"
                        and not cell.has_style
                    ):
                        continue

                    # Handle merged cells
                    if (row, col) in merged_cells:
                        top_left_row, top_left_col = merged_cells[(row, col)]

                        # If this is not the top-left cell of the merged range, skip it
                        if row != top_left_row or col != top_left_col:
                            continue

                    # Convert cell to column letter and row number (e.g., 'A1')
                    col_letter = openpyxl.utils.cell.get_column_letter(col)
                    cell_addr = f"{col_letter}{row}"

                    # Determine cell type
                    cell_type = "empty"
                    if cell.data_type == "f":
                        cell_type = "formula"
                    elif cell.data_type == "s":
                        cell_type = "string"
                    elif cell.data_type == "n":
                        cell_type = "number"
                    elif cell.data_type == "d":
                        cell_type = "date"
                    elif cell.data_type == "b":
                        cell_type = "boolean"

                    # Extract cell format
                    cell_format = "General"  # Default format
                    if cell.number_format:
                        cell_format = cell.number_format

                    # Create cell dictionary
                    cell_dict = {
                        "value": cell.value,
                        "type": cell_type,
                        "format": cell_format,
                    }

                    # Add formula if present
                    if cell_type == "formula":
                        cell_dict["formula"] = cell.value

                    # Add merged cell information if applicable
                    if (row, col) in merged_cells:
                        min_row = merged_cells[(row, col)][0]
                        min_col = merged_cells[(row, col)][1]
                        merged_range = next(
                            r
                            for r in merged_ranges
                            if r.min_row == min_row and r.min_col == min_col
                        )
                        cell_dict["merged"] = {
                            "min_row": merged_range.min_row,
                            "min_col": merged_range.min_col,
                            "max_row": merged_range.max_row,
                            "max_col": merged_range.max_col,
                            "range": f"{openpyxl.utils.cell.get_column_letter(merged_range.min_col)}{merged_range.min_row}:"
                            f"{openpyxl.utils.cell.get_column_letter(merged_range.max_col)}{merged_range.max_row}",
                        }

                    sheet_dict[cell_addr] = cell_dict

            workbook_dict[sheet_name] = sheet_dict

        logger.info(f"Loaded workbook with {len(workbook_dict)} sheets")
        return workbook_dict

    except InvalidFileException:
        logger.error(f"Invalid Excel file: {file_path}")
        raise ValueError(f"Invalid Excel file: {file_path}")
    except Exception as e:
        logger.error(f"Error loading workbook: {str(e)}")
        raise


def detect_pivot_tables(
    workbook: Dict[str, Dict[str, Any]],
) -> Dict[str, List[Dict[str, Any]]]:
    """
    Detect pivot tables in a workbook

    Args:
        workbook: Dictionary representation of the workbook

    Returns:
        Dictionary mapping sheet names to lists of pivot table information
    """
    # Implementation would detect pivot tables
    # For now, return an empty dictionary
    return {}


def detect_charts(
    workbook: Dict[str, Dict[str, Any]],
) -> Dict[str, List[Dict[str, Any]]]:
    """
    Detect charts in a workbook

    Args:
        workbook: Dictionary representation of the workbook

    Returns:
        Dictionary mapping sheet names to lists of chart information
    """
    # Implementation would detect charts
    # For now, return an empty dictionary
    return {}


def detect_data_validation(
    workbook: Dict[str, Dict[str, Any]],
) -> Dict[str, Dict[str, Dict[str, Any]]]:
    """
    Detect data validation rules in a workbook

    Args:
        workbook: Dictionary representation of the workbook

    Returns:
        Dictionary mapping sheet names to dictionaries of cell addresses to validation rules
    """
    # Implementation would detect data validation
    # For now, return an empty dictionary
    return {}
