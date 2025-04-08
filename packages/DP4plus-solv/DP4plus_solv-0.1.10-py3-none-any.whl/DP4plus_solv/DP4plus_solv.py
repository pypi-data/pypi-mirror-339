# -*- coding: utf-8 -*-
"""
Created on Sep 04 2024
@author: María M. Zanardi & Ariel M. Sarotti

-----------------------------------------------------------------------------------
-----------------------------------------------------------------------------------
          This program lets you to perform solvent dependent DP4+ calculation automatically.
The program allows the use of H and C data. Although it can give partial results
using some subset of data, it is recommended to use the complete data set.

*Conditions for its use:

        You must place in a folder: all the outputs containing NMR and Single Point
calculations, corresponding to all posible conformations and isomers under
analysis (in format *.log or *.out). The names of the file will necessarily
be coded as IsomerNumber_*.

        Additionally the folder must also contain an excel file consigning the
experimental data an labels ordered as follows:
    -experimental chemical shifts and labels of the atoms (sheet namme = 'shifts')
-------------------------------------------------------------------------------------
-------------------------------------------------------------------------------------
"""

#__version__ = "0.1.9"
#__author__ = 'María M. Zanardi & Ariel M. Sarotti'
#import modules for calculation
import glob
import os, time
import scipy.stats as stats
import pandas as pd
import numpy as np
import shutil
from pathlib import Path
from sys import exit
from math import isnan
from sklearn import linear_model
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from pathlib import Path

#import modules for CUI
import sys, os, shutil, subprocess
from PyQt5 import QtWidgets, uic
from PyQt5.QtCore import QThread, pyqtSignal, QEventLoop

# Cargar archivo UI
ui_path = os.path.join(os.path.dirname(__file__), 'UI.ui')

# Clase para los hilos de ejecución de las funciones
class AnalysisThread(QThread):
    result_signal = pyqtSignal(str)  # Señal para enviar resultados de vuelta a la UI
    finished = pyqtSignal()

    def __init__(self, directory, excel_file, iterations, solvent, df, tms):
        super().__init__()
        self.directory = directory
        self.excel_file = excel_file
        self.iterations = iterations
        self.solvent = solvent
        self.df = df
        self.tms = tms
        #self._is_running = True  # Control de ejecución

    def run(self):
        # Lógica de análisis dentro del hilo (simulación aquí, reemplaza con tu lógica real)
        ejecutar(self.directory, self.excel_file, self.iterations, self.solvent, self.df, self.tms)
        # Cuando termine el análisis, envía el resultado de vuelta a la UI
        result = f"Analysis completed for solvent: {self.solvent}, iterations: {self.iterations}"
        self.result_signal.emit(result)

        self.finished.emit()


# Clase para los hilos de ejecución de las funciones
class ExampleThread(QThread):
    result_signal = pyqtSignal(str)  # Señal para enviar resultados de vuelta a la UI

    def __init__(self):
        super().__init__()
        self.desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        self.example_path = os.path.join(os.path.dirname(__file__), 'Example_DP4plus_solvent')

    def run(self):
         # Definir la ruta destino en el escritorio
        destination_path = os.path.join(self.desktop_path, 'Example_DP4plus_solvent')

        if os.path.exists(destination_path) :
            destination_path += str(np.random.randint(0,1000))

        # Copiar el archivo al escritorio
        shutil.copytree(self.example_path, destination_path)

        # Cuando termine el análisis, envía el resultado de vuelta a la UI
        result = f"Example created in: {destination_path}"

        self.result_signal.emit(result)

class LevelDialog(QtWidgets.QDialog):
    def __init__(self):
        super(LevelDialog, self).__init__()

        # Cargar el archivo UI del dialog
        uic.loadUi(os.path.join(os.path.dirname(__file__), "UI2.ui"), self)

        # Asumiendo que la tabla en UI2.ui tiene el objeto llamado 'tableWidget'
        self.table = self.findChild(QtWidgets.QTableWidget, 'tableWidget')

        # Para agregar los tensores del estandar de referencia'
        self.tmsH = self.findChild(QtWidgets.QDoubleSpinBox, 'tmsH')
        self.tmsC = self.findChild(QtWidgets.QDoubleSpinBox, 'tmsC')

        # Conectar los botones de aceptar y cancelar
        self.btn_accept = self.findChild(QtWidgets.QPushButton, 'btn_accept')
        self.btn_cancel = self.findChild(QtWidgets.QPushButton, 'btn_cancel')

        self.btn_accept.clicked.connect(self.accept)
        self.btn_cancel.clicked.connect(self.reject)

        # Asignar QDoubleSpinBox a cada celda de la tabla si no están en el archivo .ui
        for row in range(6):
            for col in range(3):
                spin_box = QtWidgets.QDoubleSpinBox(self)  # Crear un QDoubleSpinBox
                spin_box.setRange(-100.0, 100.0)  # Rango de valores (ajústalo según lo necesario)
                spin_box.setDecimals(4)  # Número de decimales
                spin_box.setButtonSymbols(QtWidgets.QAbstractSpinBox.NoButtons) # Eliminar las flechas (botones de incremento y decremento)
                self.table.setCellWidget(row, col, spin_box)  # Asignar el QDoubleSpinBox a la celda


    # Función para extraer los datos en un DataFrame cuando se presiona "Aceptar"
    def get_values(self):
        data = []
        for row in range(6):  # Asegúrate de que la tabla en UI2.ui tenga 6 filas y 3 columnas
            row_data = []
            for col in range(3):
                spin_box = self.table.cellWidget(row, col)  # Suponemos que cada celda tiene un QDoubleSpinBox
                if isinstance(spin_box, QtWidgets.QDoubleSpinBox):
                    row_data.append(spin_box.value())
            data.append(row_data)

        # Crear el DataFrame
        df = pd.DataFrame(data, columns=['μ', 'σ', 'ν'], index=['sC', 'sH', 'UnsC_sp2', 'UnsH_sp2', 'UnsC_sp3', 'UnsH_sp3'])
        # Obtener valores de los QSpinBox para TMS
        tms = {
            'TMS_C': self.tmsC.value(),
            'TMS_H': self.tmsH.value()}

        return df, tms



# Clase principal que maneja la UI
class SolventApp(QtWidgets.QMainWindow):
    def __init__(self):
        super(SolventApp, self).__init__()
        uic.loadUi(ui_path, self)

        self.setWindowTitle("Solvent DP4+")

        # Asignación de los widgets a las variables
        self.bt_dir = self.findChild(QtWidgets.QPushButton, 'bt_dir')
        self.bt_xls = self.findChild(QtWidgets.QPushButton, 'bt_xls')
        self.lb_dir = self.findChild(QtWidgets.QLabel, 'lb_dir')
        self.lb_xls = self.findChild(QtWidgets.QLabel, 'lb_xls')
        self.iter_num = self.findChild(QtWidgets.QSpinBox, 'iter_num')
        self.solvent = self.findChild(QtWidgets.QComboBox, 'solvent')
        self.bt_exit = self.findChild(QtWidgets.QPushButton, 'bt_exit')
        self.bt_run = self.findChild(QtWidgets.QPushButton, 'bt_run')
        self.bt_UG = self.findChild(QtWidgets.QPushButton, 'bt_UG')
        self.bt_example = self.findChild(QtWidgets.QPushButton, 'bt_example')
        self.the_lev = self.findChild(QtWidgets.QComboBox, 'the_lev')


        # Conexiones de botones a funciones
        self.bt_dir.clicked.connect(self.select_directory)
        self.bt_xls.clicked.connect(self.select_excel)
        self.bt_exit.clicked.connect(self.close)
        self.bt_run.clicked.connect(self.run_analysis)
        self.bt_UG.clicked.connect(self.open_user_guide)
        self.bt_example.clicked.connect(self.load_example)

    # Función para seleccionar el directorio
    def select_directory(self):
        directory = QtWidgets.QFileDialog.getExistingDirectory(self, "Select Directory")
        if directory:
            self.lb_dir.setText(directory)

    # Función para seleccionar el archivo Excel
    def select_excel(self):
        file, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Select Excel File", "", "Excel Files (*.xls *.xlsx)")
        if file:
            self.lb_xls.setText(file)

    # Función para ejecutar el análisis en un hilo
    def run_analysis(self):
        selected_level = self.the_lev.currentText()

        if selected_level == 'Other Level':
            dialog = LevelDialog()
            if dialog.exec_() == QtWidgets.QDialog.Accepted:
                df, tms = dialog.get_values()  # Captura los valores como DataFrame
                print(df)  # Imprime el DataFrame en la consola
                print(tms)
                QtWidgets.QMessageBox.information(self, "Data Captured", f"Data captured:\n{df.to_string()}")
        else:
            df = pd.DataFrame([(0, 1.5569, 6), (0, 0.1044, 3),
                               (-0.9201, 1.7484, 5), (0.3475, 0.1176, 4),
                               (2.9085, 1.6, 6), (-0.018, 0.1118, 3)],
                              columns=['μ', 'σ', 'ν'],
                              index=['sC', 'sH', 'UnsC_sp2', 'UnsH_sp2', 'UnsC_sp3', 'UnsH_sp3'])
            tms = {'TMS_C': 196.6095, 'TMS_H': 31.560}

        # Llamar a la función que lanza el hilo con los argumentos correctos
        directory = self.lb_dir.text()
        excel_file = self.lb_xls.text()
        iterations = self.iter_num.value()
        solvent = self.solvent.currentText()

        # Verifica que los parámetros sean válidos
        if not directory or not excel_file or iterations == 0:
            QtWidgets.QMessageBox.warning(self, "Input Error", "Please select a directory, Excel file, and set iterations.")
            return

        # Inicia el análisis en un hilo
        self.run_thread = AnalysisThread(directory, excel_file, iterations, solvent, df, tms)
        self.run_thread.result_signal.connect(self.popup_information)
        self.run_thread_a_freeze_tab()

    # Función para manejar el final del análisis
    def popup_information(self, result):
        QtWidgets.QMessageBox.information(self, 'Solvent DP4+ ',result)

    # Función para abrir la guía de usuario
    def open_user_guide(self):
        #self.popup_information(os.path.join(os.path.dirname(__file__), "UserGuide", "UserGuide.pdf"))
        guide_path = os.path.join(os.path.dirname(__file__), "UserGuide", "UserGuide.pdf")
        if os.path.exists(guide_path):
            if sys.platform == "win32":
                os.startfile(guide_path)
            else:
                subprocess.call(('xdg-open', guide_path))  # Para Linux/Unix
        else:
            QtWidgets.QMessageBox.warning(self, "Error", "User guide not found.")

    # Función para cargar un ejemplo
    def load_example(self):
        # Inicia el análisis en un hilo
        self.run_thread = ExampleThread()
        self.run_thread.result_signal.connect(self.popup_information)
        self.run_thread.start()

    def run_thread_a_freeze_tab(self):
        self.setEnabled(False)

        event_loop = QEventLoop()  # Create an event loop

        self.run_thread.finished.connect(event_loop.quit) #connect the event loop quit signal to the thread finished signal
        self.run_thread.start() # Start the thread

        event_loop.exec_()  # Start the event loop
        time.sleep(1)
        self.setEnabled(True) # Enable the tab after the thread finishes

#Function definitions for the calculation
def ejecutar(directory, excel_file, iterations, solvent, df, tms):
    os.chdir(directory)
    isomer_list, cant_comp = isomer_count()
    d_exp_C, wtl_C_all, d_exp_H, wtl_H_all = data_sheet(excel_file, cant_comp)

    isomers_tensors = pd.DataFrame()
    tens_matrix_all_isomers = []

    for n_isom, isom in enumerate(isomer_list):

        conformers = glob.glob(f'{isom}*.log') +  glob.glob(f'{isom}*.out')
        conformers.sort() ##RG
        conformers.sort(key=lambda s: len(s)) #RG

        wtl_C = label_check(wtl_C_all, n_isom+1) #labels verifier
        wtl_H = label_check(wtl_H_all, n_isom+1)

        tens_all_conf = pd.DataFrame()
        energies_PCM=[]

        for n,conf in enumerate(conformers):
            '''Both energies and the tensors are extracted from each conformer.
            Both energies could be in one or separeted output files'''
            lines, energy_PCM, tensors = get_information(conf)
            if energy_PCM != 0:
                energies_PCM.append(energy_PCM)
            tens_all_conf = pd.concat([tens_all_conf,tensors],axis=1)

        '''Obtaining a vector with an applied filter, Boltzmann-weighted corrected tensors'''
        e_relativas = relative_energies(energies_PCM)
        tensores_ponderados = Boltzman(tens_all_conf, e_relativas)
        tens_matrix_all_isomers.append(tensores_ponderados)

    DP4_results = DP4_standard_calculation(cant_comp, tens_matrix_all_isomers, wtl_C_all, wtl_H_all, d_exp_C, d_exp_H, df, tms).T
    DP4_plus_solv = DP4_solvent_calculation(cant_comp, tens_matrix_all_isomers, wtl_C_all, wtl_H_all, d_exp_C, d_exp_H, solvent, iterations, df, tms)


    "Data procesing to print data to an excel file"
    isomers = isomers_names(isomer_list, cant_comp)

    DP4_types_names = ['sDP4_H', 'sDP4_C', 'sDP4_Full', 'UsDP4_H', 'UsDP4_C', 'UsDP4_Full',
                 'DP4_H', 'DP4_C', 'DP4_Full']

    DP4_results.index = DP4_types_names
    DP4_results.columns= isomers

    DP4_plus_solv.index = DP4_types_names
    DP4_plus_solv.columns= isomers


    print('\nDone!')
    perm=False

    while not perm:
        perm=True
        try:

            directorio = os.getcwd()
            base_result_file = 'DP4+_Results'
            file_path = os.path.join(directorio, f'{base_result_file}.xlsx')
            counter = 1

            while os.path.exists(file_path):
                file_path = os.path.join(directorio, f'{base_result_file} ({counter}).xlsx')
                counter += 1


            workbook = file_path
            print(f'\n -> Writting output file DP4+_Results({counter}).xlsx in {directorio:s}.')
            with open(workbook, 'w') as f:
                f.write('output file')
        except:
            print("   -> Can't write. Please choose another directory for output file.")
            perm=False


    #label_columns = [n+1 for n in range(matrix_tensors.shape[1])]
    with pd.ExcelWriter(workbook) as writer:
        DP4_results.to_excel(writer,sheet_name='CDCl3', index=True,float_format="%.2f")
        DP4_plus_solv.to_excel(writer,sheet_name=f'{solvent}', index=True,float_format="%.2f")

    modify_report(workbook, solvent)

def isomer_count():
    '''Determine the amount of isomeric candidates to be evaluated
    The files must be named by: isomerID_ * .log
    Funtion globals var "isomer_list" '''
    global isomer_list
    files= glob.glob('*.log') + glob.glob('*.out')
    isomer_list =[]
    for file in files:
        if file.split('_',1)[0] not in isomer_list:
            isomer_list.append(file.split('_',1)[0])
        else:
            continue
    isomer_list.sort() ##RG
    isomer_list.sort(key=lambda s: len(s)) #RG
    return isomer_list, len(isomer_list)

def data_sheet(open_file, cant_comp):
    'Allows you to read Excel with the label, interactively \ n'
    print(f' -> Using {open_file:s} as excel file with experimental data and labels.')
    df = pd.read_excel(open_file, sheet_name='shifts',engine='openpyxl')
    data = np.array(df[df['nuclei'].isna() == False])
    shifts = data
    d_exp_C = np.array([shifts[i][1:4] for i in range(shifts.shape[0]) if shifts[i][0] in 'cC'])
    d_exp_H = np.array([shifts[i][1:4] for i in range(shifts.shape[0]) if shifts[i][0] in 'hH'])
    if shifts.shape[1] < 7:
        wtl_C = np.array([shifts[i][4:6] for i in range(shifts.shape[0]) if shifts[i][0] in 'cC'])
        wtl_H = np.array([shifts[i][4:6] for i in range(shifts.shape[0]) if shifts[i][0] in 'hH'])
    else:
        for i in range(cant_comp):
            end_label = (cant_comp *3) + 4
            wtl_C = np.array([shifts[i][4:end_label] for i in range(shifts.shape[0]) if shifts[i][0] in 'cC'])
            wtl_H = np.array([shifts[i][4:end_label] for i in range(shifts.shape[0]) if shifts[i][0] in 'hH'])
    return d_exp_C, wtl_C, d_exp_H, wtl_H

def label_check(wtl, isom):
    '''Change de labeling if required'''
    try:
        if len(wtl[0]) < 4:
            return wtl
        else:
            start = (isom -1) *3
            end = start + 3
            return wtl[:, start:end]
    except:
        return wtl

def get_information(file):
    '''Recive an output of gaussian 09 (.log o .out) and return 4 elements:
    The txt file in lines, SCF energy at RmPW1PW91 and RB3LYP levels,
    and a DataFrame of tensors,
    '''
    tensors=[]
    energy_PCM= 0
    with open (file,'rt') as f:
        lines=f.readlines()
        for i, line in enumerate(lines):
            if "SCF Done:" and "E(RmPW1PW91)" in line:
                energy_PCM=float(line.split()[4])
            if "Isotropic = " in line:
                tensors.append(float(line.split()[4]))

    return lines, energy_PCM, pd.DataFrame(tensors)

def relative_energies(energies):
    '''Receive a list with the energies of all the conformers and get the energy
    relative to the minimum'''
    energ = np.array(energies)
    energ *= 627.5095
    mas_estable = energ.min()
    e_relativas = energ - mas_estable
    return e_relativas

def Boltzman(tens_all_conf, e_relativas):
    '''Once all the tensors  have been extracted and corrected, this function
     performs the weighted averages according to the Boltzman probability and Boltzman factor.
     Its parameters require a DataFrame with the tensors of C and H
     a list of those relative energies and the boltzman factor (Fb).
    '''
    P_Boltz = (np.exp((-e_relativas*4.18)/2.5))
    contribucion = P_Boltz / P_Boltz.sum()
    tensores_ponderados = pd.DataFrame(tens_all_conf) * contribucion
    tensores_ponderados = tensores_ponderados.sum(axis=1)

    return tensores_ponderados

def DP4_standard_calculation(cant_comp, tens_matrix_all_isomers, wtl_C_all, wtl_H_all, d_exp_C, d_exp_H, df, tms):
    UnS_shifts_C = pd.DataFrame()
    UnS_shifts_H = pd.DataFrame()
    isomers_shifts_C = pd.DataFrame()
    isomers_shifts_H = pd.DataFrame()

    for j in range(cant_comp):
        wtl_C = label_check(wtl_C_all, j+1) #labels verifier
        wtl_H = label_check(wtl_H_all, j+1)
        isom_tens = tens_matrix_all_isomers[j] #.iloc[i]
        tens_C, tens_H = tens_ordenados(isom_tens, wtl_C, wtl_H)

        '''Diastereotopic order'''
        tens_C, d_exp_C = diasterotopics(tens_C, d_exp_C)
        tens_H, d_exp_H = diasterotopics(tens_H, d_exp_H)
        '''Chemical shifts calculation and Scaling procedure '''

        Unscaled_shift_C, Unscaled_shift_H = d_calculation(tens_C, tens_H, tms)
        Scaled_shift_C, exp_C = escalado_CyH(Unscaled_shift_C, d_exp_C)
        Scaled_shift_H, exp_H = escalado_CyH(Unscaled_shift_H, d_exp_H)

        '''Once the corrected tensors and scaled chemical shifts of an isomer
        are obtained, they are added to the DataFrame of all the isomers that
        will be correlated with the experimental data '''

        UnS_shifts_C = pd.concat([UnS_shifts_C, Unscaled_shift_C ],axis=1)
        UnS_shifts_H = pd.concat([UnS_shifts_H, Unscaled_shift_H ],axis=1)

        isomers_shifts_C = pd.concat([isomers_shifts_C, Scaled_shift_C ],axis=1)
        isomers_shifts_H = pd.concat([isomers_shifts_H, Scaled_shift_H ],axis=1)

    '''Unscaled Error calculation'''
    UscEC = pd.DataFrame(error_calculator(UnS_shifts_C, exp_C))
    UscEH = pd.DataFrame(error_calculator(UnS_shifts_H, exp_H))

    '''Sp2 and Sp3 classifier'''
    UscEC_sp2, UscEC_sp3 = hibridization_classifier(UscEC, d_exp_C)
    UscEH_sp2, UscEH_sp3 = hibridization_classifier(UscEH, d_exp_H)

    '''Calculation of the scaled H and C errors'''
    scEC = error_calculator(isomers_shifts_C, exp_C)
    scEH = error_calculator(isomers_shifts_H, exp_H)

    errors_list = [scEC, scEH, UscEC_sp2, UscEH_sp2, UscEC_sp3, UscEH_sp3]

    '''Calculation of the partial or complete DP4_plus multiensambles'''
    DP4_results = DP4_plus_calculator(errors_list, cant_comp, df)
    return DP4_results

def diasterotopics(tens_C, d_exp_C):
    '''If there are diasterotopic nuclei, the smallest value of the exp data
    will be matched with the smallest value of the calculations for each isomer.
    Returns the same variables corrected'''

    tens_C2 = tens_C.copy()
    d_exp_C2 = d_exp_C.copy()

    for i in range(d_exp_C.shape[0]-1):
        if d_exp_C2[i][2]!='nan':
            for j in range(i, d_exp_C.shape[0]-1):
                if d_exp_C2[i][2] == d_exp_C2[j+1][2]:
                    d_exp_C[i][1] = max(d_exp_C2[i][1], d_exp_C2[j+1][1])
                    d_exp_C[j+1][1] = min(d_exp_C2[i][1], d_exp_C2[j+1][1])
                    tens_C[i] = min(tens_C2[i], tens_C2[j+1])
                    tens_C[j+1]= max(tens_C2[i], tens_C2[j+1])
                else:
                    continue
        else:
            continue
    return tens_C, d_exp_C

def tens_ordenados(tensors_indexed_C, wtl_C, wtl_H):
    ''' To operate you need the tensors indexed in Gaussian order and the
    labels of the compound in np.array format (this should have only 3 columns)
    The result is a list with the tensors ordered according to the inserted
    label'''
    tens_C = np.zeros(wtl_C.shape)
    tens_H = np.zeros(wtl_H.shape)
    for i in range (wtl_C.shape[0]):
        for j in range (3):
            if not isnan(wtl_C[i,j]):
                index = int(wtl_C[i,j])
                tens_C[i,j] = tensors_indexed_C[index-1]
            else:
                tens_C[i,j]=float('nan')
    for i in range (wtl_H.shape[0]):
        for j in range (3):
            if not isnan(wtl_H[i,j]):
                index = int(wtl_H[i,j])
                tens_H[i,j] = tensors_indexed_C[index-1]
            else:
                tens_H[i,j]=float('nan')
    tens_C = pd.DataFrame(tens_C)
    tens_C = tens_C.mean(axis=1)

    tens_H = pd.DataFrame(tens_H)
    tens_H = tens_H.mean(axis=1)
    return tens_C, tens_H

def d_calculation(tensors_C, tensors_H, tms):
    '''It use the tensors for the calculation of the chemical shifts using
    TMS as a reference standard'''
    TMS_C = tms['TMS_C']
    TMS_H = tms['TMS_H']
    Unscaled_shift_C = pd.DataFrame(TMS_C - tensors_C)
    Unscaled_shift_H = pd.DataFrame(TMS_H - tensors_H)
    return Unscaled_shift_C, Unscaled_shift_H

def escalado_CyH(Unscaled_shift, exp):
    '''Performs C and H scaling.
     You will also order interchangeable H or C but requiered that the
     experimental chemical shifts to exchange are labeled with the same character'''
    shifts = []
    exchange = []

    if exp.shape[0] != 0:
        for i in range(exp.shape[0]):
            shifts.append(exp[i][1])
            exchange.append(exp[i][2])
    else:
        return pd.DataFrame(exp), pd.DataFrame(exp)

    UnsC_d = Unscaled_shift.copy()
    indices_intercambiables = [i for i in range(len(exchange)) if exchange[i] ==1]
    for e in indices_intercambiables:
        UnsC_d[e] = max(Unscaled_shift[e], Unscaled_shift[e+1])
        UnsC_d[e+1] = min(Unscaled_shift[e], Unscaled_shift[e+1])


    shifts = np.array(shifts).reshape(-1,1)
    UnsC_d = np.array(UnsC_d).reshape(-1,1)
    regresion = linear_model.LinearRegression()

    if len(shifts) == 1:
        Scaled_shift = shifts
    else:
        regresion.fit(shifts, UnsC_d)

        m = regresion.coef_
        b = regresion.intercept_

        Scaled_shift = (UnsC_d - b) / m
    Scaled_shift = pd.DataFrame(Scaled_shift)

    shifts = pd.DataFrame(shifts)
    return Scaled_shift, shifts

def error_calculator(calc, exp):
    return np.array(calc - exp)

def hibridization_classifier(err, d_exp):
    UscE_sp2 = pd.DataFrame()
    UscE_sp3 = pd.DataFrame()
    err = np.array(err)
    for i in range(d_exp.shape[0]):
        if d_exp[i][0] == 1:
            UscE_sp2 = pd.concat([UscE_sp2, pd.DataFrame(err[i])], axis = 1)
        else:
            UscE_sp3 =pd.concat([UscE_sp3, pd.DataFrame(err[i])], axis = 1)
    return UscE_sp2.transpose(), UscE_sp3.transpose()

def probability(error, mu, sigma, nu):
        t_dist = stats.t(nu)
        prob = 1 - t_dist.cdf(np.abs(error - mu) / sigma)
        return prob

def DP4_plus_calculator(errors_list, cant_comp, df):
    '''Function that calculate the DP4 probability'''
    #errors_list = [scEC, scEH, UscEC_sp2, UscEH_sp2, UscEC_sp3, UscEH_sp3]
    '''dist_parameters = {0: (0, 1.5569, 6), 1: (0, 0.1044, 3),
                       2:(-0.9201, 1.7484, 5), 3: (0.3475, 0.1176, 4),
                       4: (2.9085, 1.6, 6), 5:(-0.018, 0.1118, 3)}'''

    #DP4_types_names = ['sDP4_C', 'sDP4_H', 'sDP4_Full', 'UsDP4_C', 'UsDP4_H', 'UsDP4_Full',
                 #'DP4_C', 'DP4_H', 'DP4_Full']
    probabilities = []
    for i, error_type in enumerate(errors_list):
        mu, sigma, nu = df.iloc[i]
        if len(error_type) > 0:
            p_e = probability(error_type, mu, sigma, nu)
        else:
            p_e = np.ones((1, cant_comp))
        p_e = np.prod(p_e,0)
        probabilities.append(p_e)

    #errors_list = [scEC, scEH, UscEC_sp2, UscEH_sp2, UscEC_sp3, UscEH_sp3]
    sDP4_H = np.array(probabilities[1])
    sDP4_C = np.array(probabilities[0])
    sDP4_Full = np.array([probabilities[0], probabilities[1]])
    UnsDP4_H = np.array([probabilities[3], probabilities[5]])
    UnsDP4_C = np.array([probabilities[2], probabilities[4]])
    UnsDP4_Full = np.array([probabilities[3], probabilities[5], probabilities[2], probabilities[4]])
    DP4_plus_H = np.array([probabilities[1], probabilities[3], probabilities[5]])
    DP4_plus_C = np.array([probabilities[0], probabilities[2], probabilities[4]])
    DP4_plus_Full = np.array(probabilities)

    DP4_types = [sDP4_H, sDP4_C, sDP4_Full, UnsDP4_H, UnsDP4_C, UnsDP4_Full,
                 DP4_plus_H, DP4_plus_C, DP4_plus_Full]
    DP4_results = pd.DataFrame()

    for i, prob in enumerate(DP4_types):
        if np.prod(prob) == 1:
            DP4 = pd.DataFrame(np.array(['-' for _ in range(cant_comp)]))
        else:
            if i <2:
                DP4 = pd.DataFrame(100*prob/sum(prob))
            else:
                DP4 = pd.DataFrame(100*np.prod(prob,0)/sum(np.prod(prob,0)))

        DP4_results = pd.concat([DP4_results, DP4], axis=1)

    return DP4_results

def DP4_solvent_calculation(cant_comp, tens_matrix_all_isomers, wtl_C_all, wtl_H_all, d_exp_C, d_exp_H, solvent, iteration, df, tms):
    DP4_sDP4_H = pd.DataFrame()
    DP4_sDP4_C = pd.DataFrame()
    DP4_sDP4_Full = pd.DataFrame()
    DP4_UnsDP4_H = pd.DataFrame()
    DP4_UnsDP4_C = pd.DataFrame()
    DP4_UnsDP4_Full = pd.DataFrame()
    DP4_plus_H = pd.DataFrame()
    DP4_plus_C = pd.DataFrame()
    DP4_plus_Full = pd.DataFrame()

    DP4_iteration =  pd.DataFrame()
    for i in range(iteration):
        d_exp_C_modified, d_exp_H_modified = experimental_modulator(d_exp_C, d_exp_H, solvent)
        DP4_iter = DP4_standard_calculation(cant_comp, tens_matrix_all_isomers, wtl_C_all, wtl_H_all, d_exp_C_modified, d_exp_H_modified, df, tms)

        DP4_sDP4_H = pd.concat([DP4_sDP4_H, DP4_iter.iloc[:, 0]], axis=1)
        DP4_sDP4_C = pd.concat([DP4_sDP4_C, DP4_iter.iloc[:, 1]], axis=1)
        DP4_sDP4_Full = pd.concat([DP4_sDP4_Full, DP4_iter.iloc[:, 2]], axis=1)
        DP4_UnsDP4_H = pd.concat([DP4_UnsDP4_H, DP4_iter.iloc[:, 3]], axis=1)
        DP4_UnsDP4_C = pd.concat([DP4_UnsDP4_C, DP4_iter.iloc[:, 4]], axis=1)
        DP4_UnsDP4_Full = pd.concat([DP4_UnsDP4_Full, DP4_iter.iloc[:, 5]], axis=1)
        DP4_plus_H = pd.concat([DP4_plus_H, DP4_iter.iloc[:, 6]], axis=1)
        DP4_plus_C = pd.concat([DP4_plus_C, DP4_iter.iloc[:, 7]], axis=1)
        DP4_plus_Full = pd.concat([DP4_plus_Full, DP4_iter.iloc[:, 8]], axis=1)

    DP4_sDP4_H = DP4_sDP4_H.mean(axis=1)
    DP4_sDP4_C = DP4_sDP4_C.mean(axis=1)
    DP4_sDP4_Full = DP4_sDP4_Full.mean(axis=1)
    DP4_UnsDP4_H = DP4_UnsDP4_H.mean(axis=1)
    DP4_UnsDP4_C = DP4_UnsDP4_C.mean(axis=1)
    DP4_UnsDP4_Full = DP4_UnsDP4_Full.mean(axis=1)
    DP4_plus_H= DP4_plus_H.mean(axis=1)
    DP4_plus_C =  DP4_plus_C.mean(axis=1)
    DP4_plus_Full = DP4_plus_Full.mean(axis=1)

    DP4_plus_solv = pd.DataFrame([DP4_sDP4_H, DP4_sDP4_C, DP4_sDP4_Full, DP4_UnsDP4_H, DP4_UnsDP4_C, DP4_UnsDP4_Full, DP4_plus_H, DP4_plus_C, DP4_plus_Full])
    return DP4_plus_solv

def isomers_names(isomer_list, cant_comp):
    isomers = []
    for isom in isomer_list:
        isomer = 'Isomer '
        for caracter in isom:
            if caracter != '_':
                isomer += caracter
            else:
                break
        isomers.append(isomer)
    return isomers

def modify_report(report, solvent):
    wb = load_workbook(report)
    sheets = ['DP4+ std', f'DP4+ solv']

    for sheet in sheets:
        ws = wb[sheet]
        ws.column_dimensions['A'].width = 20
        #ws['A2'].fill = PatternFill(start_color = '00CED1', end_color = '00CED1', fill_type = "solid")
        for i in range(2, 5):
            ws[f'A{i}'].fill = PatternFill(start_color = 'AFEEEE', end_color = 'AFEEEE', fill_type = "solid")
        for i in range(5, 8):
            ws[f'A{i}'].fill = PatternFill(start_color = 'FFE4E1', end_color = 'FFE4E1', fill_type = "solid")
        for i in range(8, 11):
            ws[f'A{i}'].fill = PatternFill(start_color = 'F5DEB3', end_color = 'F5DEB3', fill_type = "solid")

    wb.save(report)
    return

def experimental_modulator(d_exp_C, d_exp_H, solvent):
    '''solvent distributions: muC, sigmaC, muH, sigmaH'''
    datos_estadistica = {
    "Acetone": [0.523, 0.509, -0.009, 0.075],
    "Acetonitrile": [0.615, 0.465, -0.050, 0.062],
    "Benzene": [0.053, 0.392, -0.174, 0.208],
    "DCM": [0.233, 0.277, -0.023, 0.048],
    "DMSO": [-0.445, 0.534, -0.105, 0.100],
    "MeOD": [1.014, 0.574, 0.014, 0.077],
    "Pyridine": [0.894, 0.496, 0.189, 0.165],
    "H2O": [2.115, 1.595, 0.135, 0.230]}

    mu_C = datos_estadistica[solvent][0]
    sigma_C = datos_estadistica[solvent][1]
    mu_H = datos_estadistica[solvent][2]
    sigma_H = datos_estadistica[solvent][3]

    # Modificar SOLO la primera columna (índice 0) usando una distribución normal
    d_exp_C_modified = d_exp_C.copy()  # Hacer una copia para modificar
    d_exp_C_modified[:, 1] = d_exp_C[:, 1] - np.random.normal(mu_C, sigma_C, d_exp_C.shape[0])

    d_exp_H_modified = d_exp_H.copy()  # Hacer una copia para modificar
    d_exp_H_modified[:, 1] = d_exp_H[:, 1] - np.random.normal(mu_H, sigma_H, d_exp_H.shape[0])


    # Genera números aleatorios a partir de una distribución normal y sumar los valores aleatorios a cada elemento del vector
    #d_exp_C_modified = d_exp_C[:, 1] + np.random.normal(mu_C, sigma_C, d_exp_C.shape[0])
    #d_exp_H_modified = d_exp_H[:, 1]  + np.random.normal(mu_H, sigma_H, d_exp_H.shape[0])
    return d_exp_C_modified, d_exp_H_modified

#Funcion para crear acceso directo
def create_exe():
    '''Creates a direc acces executable file in the user desktop'''
    #desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    desktop = os.path.normpath(os.path.expanduser("~/Desktop"))
    dir_acc = os.path.join(desktop,'DP4plus_solv.py')

    with open (dir_acc, 'w') as file:
        file.write('# -*- coding: utf-8 -*-\n\n')
        file.write('import os, shutil\n\n')
        file.write('exe = shutil.which("DP4plus-solv")\n\n')
        file.write('os.system(exe)\n\n')

# Main
def main(args=None):
    app = QtWidgets.QApplication(sys.argv)
    window = SolventApp()
    window.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()
